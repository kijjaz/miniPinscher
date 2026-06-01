import openpyxl
import os
import pandas as pd

XLSX_FILE = "20260602 Kijjaz - miniPinscher 0.6.0.xlsx"

def verify_060():
    print(f"🧪 Verifying generated Spreadsheet: {XLSX_FILE}...")
    
    if not os.path.exists(XLSX_FILE):
        print(f"❌ Error: File {XLSX_FILE} not found!")
        return False
        
    wb = openpyxl.load_workbook(XLSX_FILE, read_only=True)
    sheetnames = wb.sheetnames
    print(f"  Loaded workbook sheet names: {sheetnames}")
    
    # 1. Verify sheet names
    expected_sheets = [
        "Formula", "IFRA_Compliance", "EU_Allergen_Labeling", 
        "DB_Standards", "DB_Naturals", "DB_Inventory", "DB_User_Materials", "DB_EU_Allergens"
    ]
    for s in expected_sheets:
        if s not in sheetnames:
            print(f"❌ Error: Missing sheet '{s}'!")
            return False
            
    print("  All 8 expected tabs are present.")

    # 2. Verify formulas in cell A1
    wb_full = openpyxl.load_workbook(XLSX_FILE, data_only=False)
    
    ws_ifra = wb_full["IFRA_Compliance"]
    formula_ifra = ws_ifra["A1"].value
    if formula_ifra != '=IFRA_COMPLIANCE_V5(Formula!A2:B, DB_Standards!A2:C, DB_Naturals!A2:D, DB_Inventory!A2:B, DB_User_Materials!A2:E, Formula!E1)':
        print(f"❌ Error: Incorrect IFRA formula: {formula_ifra}")
        return False
    print(f"  Verified IFRA Compliance cell A1 formula: {formula_ifra}")

    ws_eu = wb_full["EU_Allergen_Labeling"]
    formula_eu = ws_eu["A1"].value
    if formula_eu != '=EU_LABELING_V5(Formula!A2:B, DB_EU_Allergens!A2:B, DB_Naturals!A2:D, DB_Inventory!A2:B, DB_User_Materials!A2:E, "Leave-on", Formula!E1)':
        print(f"❌ Error: Incorrect EU formula: {formula_eu}")
        return False
    print(f"  Verified EU Allergen Labeling cell A1 formula: {formula_eu}")

    # 3. Verify Database Tab row counts dynamically against release CSVs
    db_files = {
        "DB_Standards": "DB_Standards.csv",
        "DB_Naturals": "DB_Naturals.csv",
        "DB_Inventory": "DB_Inventory.csv",
        "DB_User_Materials": "DB_User_Materials.csv",
        "DB_EU_Allergens": "DB_EU_Allergens.csv"
    }
    
    release_dir = "google_sheets_release"
    for tab, csv_filename in db_files.items():
        csv_path = os.path.join(release_dir, csv_filename)
        if not os.path.exists(csv_path):
            print(f"❌ Error: Source CSV {csv_path} not found for tab '{tab}'!")
            return False
        
        # Count rows in CSV
        try:
            df_csv = os.popen(f"wc -l < {csv_path}").read().strip()
            # Alternative: read using pandas for precision
            df_csv = pd.read_csv(csv_path)
            expected_rows = len(df_csv) + 1 # +1 for header
        except Exception as e:
            print(f"❌ Error reading CSV {csv_path}: {e}")
            return False

        ws = wb_full[tab]
        row_count = ws.max_row
        if row_count != expected_rows:
            print(f"❌ Error: Tab '{tab}' row count is {row_count}, expected {expected_rows} (from {csv_filename})!")
            return False
        print(f"  Verified tab '{tab}' row count matches CSV exactly: {row_count} rows.")

    print(f"\n✅ Verification Successful: {XLSX_FILE} is 100% complete, fully validated, and structurally healthy!")
    return True

if __name__ == "__main__":
    verify_060()
