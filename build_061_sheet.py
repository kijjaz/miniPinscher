import openpyxl
import pandas as pd
import os

RELEASE_DIR = "google_sheets_release"
OUTPUT_XLSX_ROOT = "20260602 Kijjaz - miniPinscher 0.6.1.xlsx"
OUTPUT_XLSX_RELEASE = os.path.join(RELEASE_DIR, "20260602 Kijjaz - miniPinscher 0.6.1.xlsx")

def build_061_sheet():
    print("🚀 Generating miniPinscher Spreadsheet Version 0.6.1...")
    
    # 1. Create a fresh workbook
    wb = openpyxl.Workbook()
    
    # 2. Add 'Formula' Sheet (Workplace)
    ws_formula = wb.active
    ws_formula.title = "Formula"
    ws_formula.append(["Ingredient", "Amount"])
    
    # Pre-populate with a working test formula
    test_formula = [
        ["Bergamot oil expressed", 2.5],
        ["Lemon oil cold pressed", 1.0],
        ["Alpha Damascone", 0.01],
        ["Beta Damascone", 0.02],
        ["Hedione", 10.0],
        ["DPG", 86.47]
    ]
    for row in test_formula:
        ws_formula.append(row)
        
    # Write Finished Dosage Configuration Cell
    ws_formula["D1"] = "Finished Dosage (%)"
    ws_formula["E1"] = 20
    print("  Created 'Formula' workspace tab with Finished Dosage (%) pre-set in cell E1.")

    # 3. Add 'IFRA_Compliance' Sheet
    ws_ifra = wb.create_sheet("IFRA_Compliance")
    ws_ifra["A1"] = '=IFRA_COMPLIANCE_V5(Formula!A2:B, DB_Standards!A2:C, DB_Naturals!A2:D, DB_Inventory!A2:B, DB_User_Materials!A2:E, Formula!E1)'
    print("  Created 'IFRA_Compliance' custom calculation tab.")

    # 4. Add 'EU_Allergen_Labeling' Sheet
    ws_eu = wb.create_sheet("EU_Allergen_Labeling")
    ws_eu["A1"] = '=EU_LABELING_V5(Formula!A2:B, DB_EU_Allergens!A2:B, DB_Naturals!A2:D, DB_Inventory!A2:B, DB_User_Materials!A2:E, "Leave-on", Formula!E1)'
    print("  Created 'EU_Allergen_Labeling' custom calculation tab.")

    # 5. Populate Database Tabs from CSVs
    db_files = {
        "DB_Standards": "DB_Standards.csv",
        "DB_Naturals": "DB_Naturals.csv",
        "DB_Inventory": "DB_Inventory.csv",
        "DB_User_Materials": "DB_User_Materials.csv",
        "DB_EU_Allergens": "DB_EU_Allergens.csv"
    }
    
    for tab_name, csv_file in db_files.items():
        csv_p = os.path.join(RELEASE_DIR, csv_file)
        if not os.path.exists(csv_p):
            print(f"❌ Error: Required database file {csv_p} is missing!")
            return False
            
        print(f"  Populating tab '{tab_name}' from {csv_file}...")
        df = pd.read_csv(csv_p)
        ws_db = wb.create_sheet(tab_name)
        
        # Append headers
        ws_db.append(df.columns.tolist())
        
        # Append data rows
        for row in df.itertuples(index=False, name=None):
            row_clean = ["" if pd.isna(v) else v for v in row]
            ws_db.append(row_clean)
            
        print(f"    Tab '{tab_name}' complete ({len(df)} rows loaded).")

    # 6. Save Excel file to both root and release folders
    wb.save(OUTPUT_XLSX_ROOT)
    wb.save(OUTPUT_XLSX_RELEASE)
    print(f"\n✅ Excel Workbook v0.6.1 saved successfully to:\n  - {OUTPUT_XLSX_ROOT}\n  - {OUTPUT_XLSX_RELEASE}")
    
    return True

if __name__ == "__main__":
    build_061_sheet()
